from T2.DIP import *

###Sabiendo que es s,S
start_time = time.time()
#(s,S)
G = collections.defaultdict(dict)
C = collections.defaultdict(dict)

for y in range(Q+1):
  C[T+1][y] = 0

S = {}
s = {}
pi = {}
for t in reversed(range(1,T+1)):
  m = semana_mes[t]
  probD = meses[m]
  for y in range(Q+1):
      loss = c * y
      for d in probD.keys():
          loss += q*max(d-y,0)*probD[d]
          loss += h*max(y-d,0)*probD[d]
          loss += C[t+1][max(y-d,0)]*probD[d]
      G[t][y] = loss
  S[t] = min(G[t], key=G[t].get) #argumento
  s[t] = max(s for s in range(S[t]+1) if G[t][s] >= G[t][S[t]]+K) #argumento
  for y in range(Q+1):
    cost = (G[t][S[t]] + K - c*y if y <= s[t] else G[t][y] - c * y)
    C[t][y] = cost

print("Tiempo: {} segundos.".format(time.time() - start_time))


for t in range(1,T+1):
  pi[t] = (s[t],S[t])
pi







from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = "t"
ws["B1"] = "s"
ws["C1"] = "S"
for t in range(1,T+1):
    ws["A"+str(t+1)] = t
    ws["B"+str(t+1)] = pi[t][0]
    ws["C"+str(t+1)] = pi[t][1]


# wb.save("politica.xlsx")
